from flask import Flask, render_template, request, jsonify, make_response
import openai
import json
import csv
import io
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Configure OpenAI - Handle environment variables safely
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # Continue without dotenv if not available

# Configure OpenAI API key
openai_api_key = os.getenv('OPENAI_API_KEY')
if openai_api_key:
    openai.api_key = openai_api_key
else:
    print("Warning: OPENAI_API_KEY not found in environment variables")

# Sample hardcoded tasks for demo
SAMPLE_TASKS = [
    "need to call client about the website redesign project they mentioned last week also check if they want mobile app too",
    "fix the login bug that users are reporting on the dashboard page asap before end of week",
    "review sarah's code for the payment integration feature and provide feedback by tomorrow",
    "schedule team meeting to discuss Q4 roadmap and budget planning stuff",
    "update documentation for the new API endpoints we released last month",
    "research competitor pricing models for our new subscription service launch",
    "follow up with legal team about data privacy compliance requirements for EU users",
    "optimize database queries causing slow loading times on reports page",
    "create wireframes for the new onboarding flow design mockups",
    "prepare presentation slides for stakeholder meeting next tuesday about project status"
]

def process_tasks_with_ai(tasks):
    """Process tasks using OpenAI GPT-4 to summarize, tag, and prioritize"""
    
    # Check if API key is available
    if not openai_api_key:
        print("No OpenAI API key found - using fallback processing")
        return get_fallback_processing(tasks)
    
    prompt = f"""
You are an expert project manager assistant. I need you to process the following tasks and for each task provide:

1. A clear, concise summary (max 10 words)
2. 1-2 relevant tags from these options: #urgent, #frontend, #backend, #client, #meeting, #documentation, #bug, #research, #design, #legal, #optimization, #planning
3. Priority score from 1-5 (1=low, 5=critical/urgent)

Tasks to process:
{json.dumps(tasks, indent=2)}

Please respond with a valid JSON array where each object has this exact structure:
{{
  "original": "original task text",
  "summary": "concise summary",
  "tags": ["#tag1", "#tag2"],
  "priority": number_between_1_and_5
}}

Ensure the response is valid JSON only, no additional text.
"""
    
    try:
        # Initialize OpenAI client properly
        client = openai.OpenAI(api_key=openai_api_key)
        
        response = client.chat.completions.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": "You are a helpful project management assistant that processes tasks efficiently."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=2000
        )
        
        result = response.choices[0].message.content.strip()
        # Remove any markdown code blocks if present
        if result.startswith('```json'):
            result = result[7:]
        if result.endswith('```'):
            result = result[:-3]
        
        processed_tasks = json.loads(result)
        return processed_tasks
        
    except Exception as e:
        print(f"Error processing tasks: {str(e)}")
        return get_fallback_processing(tasks)

def get_fallback_processing(tasks):
    """Fallback processing when OpenAI is not available - Fast rule-based processing"""
    processed = []
    
    for task in tasks:
        task_lower = task.lower()
        
        # Smart summary (first 7 words or key action)
        words = task.split()
        if len(words) <= 7:
            summary = task
        else:
            # Try to extract action + object
            action_words = ['fix', 'update', 'review', 'call', 'schedule', 'create', 'prepare', 'research', 'follow up', 'optimize']
            action = next((word for word in words[:3] if word.lower() in action_words), words[0])
            summary = f"{action.title()} {' '.join(words[1:4])}"
        
        # Smart tagging based on keywords
        tags = []
        if any(word in task_lower for word in ['bug', 'error', 'issue', 'problem', 'broken']):
            tags.append('#bug')
        if any(word in task_lower for word in ['urgent', 'asap', 'critical', 'immediately', 'emergency']):
            tags.append('#urgent')
        if any(word in task_lower for word in ['client', 'customer', 'user']):
            tags.append('#client')
        if any(word in task_lower for word in ['meeting', 'schedule', 'discuss', 'call']):
            tags.append('#meeting')
        if any(word in task_lower for word in ['document', 'documentation', 'docs']):
            tags.append('#documentation')
        if any(word in task_lower for word in ['frontend', 'ui', 'interface', 'design', 'wireframe']):
            tags.append('#frontend')
        if any(word in task_lower for word in ['backend', 'database', 'server', 'api']):
            tags.append('#backend')
        if any(word in task_lower for word in ['research', 'investigate', 'analyze']):
            tags.append('#research')
        if any(word in task_lower for word in ['legal', 'compliance', 'privacy']):
            tags.append('#legal')
        if any(word in task_lower for word in ['optimize', 'performance', 'slow', 'speed']):
            tags.append('#optimization')
        
        if not tags:
            tags.append('#planning')
        
        # Keep only first 2 tags
        tags = tags[:2]
        
        # Smart priority based on keywords
        priority = 3  # default
        if any(word in task_lower for word in ['asap', 'urgent', 'critical', 'emergency', 'immediately']):
            priority = 5
        elif any(word in task_lower for word in ['bug', 'error', 'broken', 'issue']):
            priority = 4
        elif any(word in task_lower for word in ['tomorrow', 'today', 'deadline']):
            priority = 4
        elif any(word in task_lower for word in ['research', 'documentation', 'explore']):
            priority = 2
        
        processed.append({
            "original": task,
            "summary": summary[:50],  # Limit length
            "tags": tags,
            "priority": priority
        })
    
    return processed

@app.route('/')
def index():
    """Render the main page"""
    return render_template('index.html')

@app.route('/api/process-tasks', methods=['POST'])
def process_tasks():
    """Process tasks via API"""
    try:
        data = request.get_json()
        
        if 'tasks' in data:
            tasks = data['tasks']
        else:
            # Use sample tasks if none provided
            tasks = SAMPLE_TASKS
        
        # Filter out empty tasks
        tasks = [task.strip() for task in tasks if task.strip()]
        
        if not tasks:
            return jsonify({"error": "No valid tasks provided"}), 400
        
        processed_tasks = process_tasks_with_ai(tasks)
        
        return jsonify({
            "success": True,
            "processed_tasks": processed_tasks,
            "total_tasks": len(processed_tasks),
            "processed_at": datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    """Export processed tasks to beautifully formatted Excel file with color coding"""
    try:
        data = request.get_json()
        processed_tasks = data.get('tasks', [])
        
        if not processed_tasks:
            return jsonify({"error": "No tasks to export"}), 400
        
        # Create a new workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "AI Processed Tasks"
        
        # Define color schemes for priorities
        priority_colors = {
            1: {'bg': 'D4F6D4', 'text': '0D5016'},  # Light green for low priority
            2: {'bg': 'E8F5E8', 'text': '1B5E20'},  # Lighter green
            3: {'bg': 'FFF8DC', 'text': '8B6914'},  # Light yellow for medium
            4: {'bg': 'FFE4E1', 'text': 'B71C1C'},  # Light red for high
            5: {'bg': 'FFCDD2', 'text': '8B0000'}   # Darker red for urgent
        }
        
        # Define styles
        header_font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        data_font = Font(name='Segoe UI', size=10)
        data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        # Headers
        headers = ['Original Task', 'AI Summary', 'Smart Tags', 'Priority Level', 'Priority Score', 'Processed Date']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row_num, task in enumerate(processed_tasks, 2):
            priority = task.get('priority', 3)
            priority_color = priority_colors.get(priority, priority_colors[3])
            
            # Original Task (A)
            cell = ws.cell(row=row_num, column=1, value=task.get('original', ''))
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # AI Summary (B)
            cell = ws.cell(row=row_num, column=2, value=task.get('summary', ''))
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # Smart Tags (C)
            tags_text = ', '.join(task.get('tags', []))
            cell = ws.cell(row=row_num, column=3, value=tags_text)
            cell.font = Font(name='Segoe UI', size=10, italic=True, color='2563EB')
            cell.alignment = data_alignment
            cell.border = border
            
            # Priority Level (D) - Text description
            priority_levels = {1: 'Low', 2: 'Low-Medium', 3: 'Medium', 4: 'High', 5: 'Urgent'}
            priority_text = priority_levels.get(priority, 'Medium')
            cell = ws.cell(row=row_num, column=4, value=priority_text)
            cell.font = Font(name='Segoe UI', size=10, bold=True, color=priority_color['text'])
            cell.fill = PatternFill(start_color=priority_color['bg'], end_color=priority_color['bg'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # Priority Score (E) - Numeric value
            cell = ws.cell(row=row_num, column=5, value=priority)
            cell.font = Font(name='Segoe UI', size=10, bold=True, color=priority_color['text'])
            cell.fill = PatternFill(start_color=priority_color['bg'], end_color=priority_color['bg'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # Processed Date (F)
            cell = ws.cell(row=row_num, column=6, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Auto-adjust column widths
        column_widths = {
            'A': 50,  # Original Task
            'B': 35,  # AI Summary
            'C': 25,  # Smart Tags
            'D': 15,  # Priority Level
            'E': 12,  # Priority Score
            'F': 20   # Processed Date
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Set row heights
        ws.row_dimensions[1].height = 25  # Header row
        for row_num in range(2, len(processed_tasks) + 2):
            ws.row_dimensions[row_num].height = 60  # Data rows with more space
        
        # Add title and metadata
        ws.insert_rows(1)
        title_cell = ws.cell(row=1, column=1, value="🎯 Smart Task Summarizer + Tagger - AI Processed Results")
        title_cell.font = Font(name='Segoe UI', size=14, bold=True, color='1E40AF')
        ws.merge_cells('A1:F1')
        
        ws.insert_rows(2)
        subtitle_cell = ws.cell(row=2, column=1, value=f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | Total Tasks: {len(processed_tasks)}")
        subtitle_cell.font = Font(name='Segoe UI', size=10, italic=True, color='64748B')
        ws.merge_cells('A2:F2')
        
        # Add an empty row for spacing
        ws.insert_rows(3)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=AI_Processed_Tasks_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        return jsonify({"error": f"Export error: {str(e)}"}), 500

@app.route('/api/export-csv', methods=['POST'])
def export_csv():
    """Export processed tasks to CSV"""
    try:
        data = request.get_json()
        processed_tasks = data.get('tasks', [])
        
        if not processed_tasks:
            return jsonify({"error": "No tasks to export"}), 400
        
        # Create CSV content
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Original Task', 'Summary', 'Tags', 'Priority', 'Processed Date'])
        
        # Write data
        for task in processed_tasks:
            writer.writerow([
                task.get('original', ''),
                task.get('summary', ''),
                ', '.join(task.get('tags', [])),
                task.get('priority', ''),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        # Create response
        csv_content = output.getvalue()
        output.close()
        
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=processed_tasks_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/sample-tasks')
def get_sample_tasks():
    """Get sample tasks for demo"""
    return jsonify({"tasks": SAMPLE_TASKS})

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000) 